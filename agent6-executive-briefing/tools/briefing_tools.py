"""
Agent 6 — Deterministic reporting tools.
Generates executive summaries, PDFs, notification payloads, and archives.
"""

import json, uuid
from datetime import datetime
from pathlib import Path

_REPORTS_DIR = Path(__file__).resolve().parent.parent / "reports"
_REPORTS_DIR.mkdir(exist_ok=True)

_report_archive: list = []


def _ts() -> str:
    return datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")


def _fmt_cost(cost: float) -> str:
    return f"${cost:,.0f}"


def generate_executive_summary(incident_id: str, purchase_order_id: str, supplier: str,
                                 execution_status: str, estimated_cost: float,
                                 approval_status: str) -> dict:
    status_line = "successfully completed" if execution_status == "SUCCESS" else "encountered issues"
    summary = (
        f"Incident {incident_id} has been {status_line}. "
        f"A purchase order ({purchase_order_id}) was issued to {supplier} "
        f"for an estimated value of {_fmt_cost(estimated_cost)}. "
        f"Approval status: {approval_status}. "
        f"The procurement workflow executed with status: {execution_status}."
    )
    return {
        "incident_id": incident_id,
        "executive_summary": summary,
        "generated_at": _ts(),
    }


def generate_incident_timeline(incident_id: str, purchase_order_id: str, supplier: str,
                                 execution_status: str, approval_status: str) -> list:
    now = datetime.utcnow()
    return [
        {"step": 1, "event": "Incident Detected",        "status": "COMPLETED", "timestamp": _ts()},
        {"step": 2, "event": "Supply Chain Analysis",    "status": "COMPLETED", "timestamp": _ts()},
        {"step": 3, "event": "Budget Assessment",        "status": "COMPLETED", "timestamp": _ts()},
        {"step": 4, "event": "Vendor Verification",      "status": "COMPLETED", "timestamp": _ts()},
        {"step": 5, "event": "Risk Assessment",          "status": "COMPLETED", "timestamp": _ts()},
        {"step": 6, "event": f"Approval ({approval_status})", "status": approval_status, "timestamp": _ts()},
        {"step": 7, "event": f"PO Generated ({purchase_order_id})", "status": "COMPLETED", "timestamp": _ts()},
        {"step": 8, "event": f"ERP Order Created",       "status": "COMPLETED", "timestamp": _ts()},
        {"step": 9, "event": f"Procurement {execution_status}", "status": execution_status, "timestamp": _ts()},
    ]


def generate_pdf_report(incident_id: str, executive_summary: str, purchase_order_id: str,
                          supplier: str, estimated_cost: float, approval_status: str,
                          execution_status: str, timeline: list) -> dict:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import (
        BaseDocTemplate, PageTemplate, Frame, Paragraph,
        Spacer, Table, TableStyle, HRFlowable,
    )
    from reportlab.pdfgen.canvas import Canvas

    # ── Palette ──────────────────────────────────────────────────────────────
    NAVY   = colors.HexColor("#0f172a")
    BLUE   = colors.HexColor("#3b82f6")
    CYAN   = colors.HexColor("#06b6d4")
    LIGHT  = colors.HexColor("#f1f5f9")
    MID    = colors.HexColor("#e2e8f0")
    MUTED  = colors.HexColor("#64748b")
    GREEN  = colors.HexColor("#22c55e")
    ORANGE = colors.HexColor("#f59e0b")
    RED    = colors.HexColor("#ef4444")
    WHITE  = colors.white

    STATUS_COLOR = GREEN if execution_status == "SUCCESS" else RED
    APPROVAL_COLOR = GREEN if approval_status == "APPROVED" else ORANGE

    W, H = A4
    ML = MR = 2 * cm
    MT = 3.2 * cm   # leave room for header band
    MB = 2.2 * cm   # leave room for footer

    filename = f"executive_report_{incident_id}_{uuid.uuid4().hex[:6]}.pdf"
    filepath = _REPORTS_DIR / filename

    # ── Header / Footer drawn on every page ──────────────────────────────────
    def _header_footer(canvas: Canvas, doc):
        canvas.saveState()
        # Header band
        canvas.setFillColor(NAVY)
        canvas.rect(0, H - 2.4 * cm, W, 2.4 * cm, fill=1, stroke=0)
        # Gradient accent line
        canvas.setFillColor(BLUE)
        canvas.rect(0, H - 2.4 * cm, W * 0.55, 0.18 * cm, fill=1, stroke=0)
        canvas.setFillColor(CYAN)
        canvas.rect(W * 0.55, H - 2.4 * cm, W * 0.45, 0.18 * cm, fill=1, stroke=0)
        # Logo text
        canvas.setFillColor(WHITE)
        canvas.setFont("Helvetica-Bold", 13)
        canvas.drawString(ML, H - 1.55 * cm, "Enterprise AI Hub")
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.HexColor("#94a3b8"))
        canvas.drawString(ML, H - 2.0 * cm, "Executive Briefing & Reporting  ·  Agent 6")
        # Right: incident badge
        canvas.setFillColor(BLUE)
        canvas.roundRect(W - MR - 4.2 * cm, H - 2.1 * cm, 4.2 * cm, 0.9 * cm, 4, fill=1, stroke=0)
        canvas.setFillColor(WHITE)
        canvas.setFont("Helvetica-Bold", 8)
        canvas.drawCentredString(W - MR - 2.1 * cm, H - 1.65 * cm, incident_id)
        # Footer band
        canvas.setFillColor(NAVY)
        canvas.rect(0, 0, W, 1.5 * cm, fill=1, stroke=0)
        canvas.setFillColor(colors.HexColor("#94a3b8"))
        canvas.setFont("Helvetica", 7.5)
        canvas.drawString(ML, 0.55 * cm, f"Generated: {_ts()}   ·   Enterprise AI Hub — Confidential")
        canvas.drawRightString(W - MR, 0.55 * cm, f"Page {doc.page}")
        canvas.restoreState()

    # ── Styles ────────────────────────────────────────────────────────────────
    def _style(name, **kw):
        base = kw.pop("parent", "Normal")
        defaults = dict(fontName="Helvetica", fontSize=9, leading=14, textColor=colors.HexColor("#1e293b"))
        defaults.update(kw)
        return ParagraphStyle(name, **defaults)

    sTitle   = _style("sTitle",   fontName="Helvetica-Bold", fontSize=22, leading=28, textColor=NAVY)
    sSub     = _style("sSub",     fontName="Helvetica",      fontSize=10, leading=14, textColor=MUTED)
    sSecHdr  = _style("sSecHdr",  fontName="Helvetica-Bold", fontSize=10, leading=14, textColor=WHITE)
    sBody    = _style("sBody",    fontName="Helvetica",      fontSize=9,  leading=14)
    sLabel   = _style("sLabel",   fontName="Helvetica-Bold", fontSize=8,  leading=12, textColor=MUTED)
    sVal     = _style("sVal",     fontName="Helvetica-Bold", fontSize=9,  leading=13, textColor=NAVY)

    def section_header(text, color=NAVY):
        """Colored band with white label."""
        data = [[Paragraph(f"<b>{text}</b>", sSecHdr)]]
        t = Table(data, colWidths=[W - ML - MR])
        t.setStyle(TableStyle([
            ("BACKGROUND",  (0, 0), (-1, -1), color),
            ("TOPPADDING",  (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ]))
        return t

    def kv_table(rows, col1=5*cm, col2=None):
        col2 = col2 or (W - ML - MR - col1)
        data = [[Paragraph(k, sLabel), Paragraph(v, sVal)] for k, v in rows]
        t = Table(data, colWidths=[col1, col2])
        t.setStyle(TableStyle([
            ("ROWBACKGROUNDS", (0, 0), (-1, -1), [LIGHT, WHITE]),
            ("LEFTPADDING",   (0, 0), (-1, -1), 10),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 10),
            ("TOPPADDING",    (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("LINEBELOW",     (0, 0), (-1, -1), 0.3, MID),
        ]))
        return t

    # ── Build story ───────────────────────────────────────────────────────────
    story = []

    # Cover title block
    story.append(Spacer(1, 0.6 * cm))
    story.append(Paragraph("Executive Briefing Report", sTitle))
    story.append(Spacer(1, 0.2 * cm))
    story.append(Paragraph(f"Incident <b>{incident_id}</b>  ·  Procurement Workflow Summary", sSub))
    story.append(Spacer(1, 0.3 * cm))
    story.append(HRFlowable(width="100%", thickness=1.5, color=BLUE, spaceAfter=0.4 * cm))

    # KPI strip (4 boxes in one row)
    kpi_data = [[
        Paragraph(f"<font color='#64748b' size=7>PURCHASE ORDER</font><br/><b>{purchase_order_id}</b>", sBody),
        Paragraph(f"<font color='#64748b' size=7>SUPPLIER</font><br/><b>{supplier}</b>", sBody),
        Paragraph(f"<font color='#64748b' size=7>ESTIMATED COST</font><br/><b>{_fmt_cost(estimated_cost)}</b>", sBody),
        Paragraph(f"<font color='#64748b' size=7>EXECUTION</font><br/><b>{execution_status}</b>", sBody),
    ]]
    col_w = (W - ML - MR) / 4
    kpi = Table(kpi_data, colWidths=[col_w] * 4)
    kpi.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), LIGHT),
        ("TOPPADDING",    (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("LEFTPADDING",   (0, 0), (-1, -1), 12),
        ("LINEAFTER",     (0, 0), (2, 0),   0.5, MID),
        ("ROUNDEDCORNERS", [4]),
    ]))
    story.append(kpi)
    story.append(Spacer(1, 0.5 * cm))

    # Executive Summary
    story.append(section_header("01  EXECUTIVE SUMMARY", NAVY))
    story.append(Spacer(1, 0.2 * cm))
    story.append(Paragraph(executive_summary, sBody))
    story.append(Spacer(1, 0.5 * cm))

    # Procurement Details
    story.append(section_header("02  PROCUREMENT DETAILS", BLUE))
    story.append(kv_table([
        ("Purchase Order ID", purchase_order_id),
        ("Supplier",          supplier),
        ("Estimated Cost",    _fmt_cost(estimated_cost)),
        ("Approval Status",   approval_status),
        ("Execution Status",  execution_status),
    ]))
    story.append(Spacer(1, 0.5 * cm))

    # Status Summary
    story.append(section_header("03  STATUS OVERVIEW", colors.HexColor("#0f766e")))
    story.append(Spacer(1, 0.2 * cm))
    status_rows = [[
        Paragraph("<font color='#64748b' size=7>APPROVAL</font>", sBody),
        Paragraph("<font color='#64748b' size=7>EXECUTION</font>", sBody),
        Paragraph("<font color='#64748b' size=7>DASHBOARD</font>", sBody),
    ], [
        Paragraph(f"<b>{approval_status}</b>", ParagraphStyle("x", fontName="Helvetica-Bold", fontSize=11,
                   textColor=APPROVAL_COLOR, leading=16)),
        Paragraph(f"<b>{execution_status}</b>", ParagraphStyle("y", fontName="Helvetica-Bold", fontSize=11,
                   textColor=STATUS_COLOR, leading=16)),
        Paragraph("<b>UPDATED</b>", ParagraphStyle("z", fontName="Helvetica-Bold", fontSize=11,
                   textColor=BLUE, leading=16)),
    ]]
    col_w3 = (W - ML - MR) / 3
    st = Table(status_rows, colWidths=[col_w3] * 3)
    st.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), LIGHT),
        ("TOPPADDING",    (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("LEFTPADDING",   (0, 0), (-1, -1), 14),
        ("LINEAFTER",     (0, 0), (1, -1),  0.5, MID),
    ]))
    story.append(st)
    story.append(Spacer(1, 0.5 * cm))

    # Incident Timeline
    story.append(section_header("04  INCIDENT TIMELINE", colors.HexColor("#7c3aed")))
    tl_header = [["#", "Event", "Status", "Timestamp"]]
    tl_rows = [[str(r["step"]), r["event"], r["status"], r["timestamp"]] for r in timeline]
    tl_data = tl_header + tl_rows
    tl = Table(tl_data, colWidths=[1.2*cm, 6.5*cm, 3*cm, 5.3*cm])
    tl_style = [
        ("BACKGROUND",    (0, 0), (-1, 0),  colors.HexColor("#7c3aed")),
        ("TEXTCOLOR",     (0, 0), (-1, 0),  WHITE),
        ("FONTNAME",      (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, 0),  8),
        ("ROWBACKGROUNDS",(0, 1), (-1, -1), [LIGHT, WHITE]),
        ("FONTSIZE",      (0, 1), (-1, -1), 8),
        ("GRID",          (0, 0), (-1, -1), 0.3, MID),
        ("TOPPADDING",    (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING",   (0, 0), (-1, -1), 8),
    ]
    # Color status cells
    for i, row in enumerate(tl_rows, start=1):
        s = row[2]
        c = GREEN if s == "COMPLETED" else (ORANGE if s == "APPROVED" else RED)
        tl_style.append(("TEXTCOLOR", (2, i), (2, i), c))
        tl_style.append(("FONTNAME",  (2, i), (2, i), "Helvetica-Bold"))
    tl.setStyle(TableStyle(tl_style))
    story.append(tl)
    story.append(Spacer(1, 0.5 * cm))

    # Confidentiality notice
    story.append(HRFlowable(width="100%", thickness=0.5, color=MID, spaceBefore=0.2*cm, spaceAfter=0.2*cm))
    story.append(Paragraph(
        "<font color='#94a3b8' size=7>CONFIDENTIAL — This report is intended solely for executive use within "
        "Enterprise AI Hub. Do not distribute without authorisation.</font>", sBody))

    # ── Assemble doc ─────────────────────────────────────────────────────────
    doc = BaseDocTemplate(
        str(filepath), pagesize=A4,
        leftMargin=ML, rightMargin=MR, topMargin=MT, bottomMargin=MB,
    )
    frame = Frame(ML, MB, W - ML - MR, H - MT - MB, id="main")
    doc.addPageTemplates([PageTemplate(id="main", frames=[frame], onPage=_header_footer)])
    doc.build(story)

    return {"report_file": filename, "report_path": str(filepath), "status": "GENERATED"}


def generate_excel_report(incident_id: str, executive_summary: str, purchase_order_id: str,
                           supplier: str, estimated_cost: float, approval_status: str,
                           execution_status: str, timeline: list) -> dict:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter

    NAVY   = "0F172A"
    BLUE   = "3B82F6"
    TEAL   = "0F766E"
    PURPLE = "7C3AED"
    LIGHT  = "F1F5F9"
    MID    = "E2E8F0"
    GREEN  = "22C55E"
    ORANGE = "F59E0B"
    RED    = "EF4444"
    WHITE  = "FFFFFF"

    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr_font(color=WHITE, sz=10, bold=True):
        return Font(name="Calibri", bold=bold, color=color, size=sz)

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def center():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    def left():
        return Alignment(horizontal="left", vertical="center", wrap_text=True)

    wb = openpyxl.Workbook()

    # ── Sheet 1: Overview ────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Overview"
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions["A"].width = 26
    ws1.column_dimensions["B"].width = 42

    # Title banner
    ws1.merge_cells("A1:B1")
    ws1["A1"] = "Enterprise AI Hub — Executive Briefing Report"
    ws1["A1"].font = Font(name="Calibri", bold=True, size=14, color=WHITE)
    ws1["A1"].fill = fill(NAVY)
    ws1["A1"].alignment = center()
    ws1.row_dimensions[1].height = 32

    ws1.merge_cells("A2:B2")
    ws1["A2"] = f"Incident: {incident_id}   ·   Generated: {_ts()}"
    ws1["A2"].font = Font(name="Calibri", size=9, color="64748B")
    ws1["A2"].fill = fill("1E293B")
    ws1["A2"].alignment = center()
    ws1.row_dimensions[2].height = 18

    ws1.append([])

    # Section header helper
    def sec(ws, row, label, color):
        ws.merge_cells(f"A{row}:B{row}")
        c = ws.cell(row=row, column=1, value=label)
        c.font = hdr_font(WHITE, 10)
        c.fill = fill(color)
        c.alignment = left()
        ws.row_dimensions[row].height = 20

    def kv(ws, row, key, val, shade=False):
        k = ws.cell(row=row, column=1, value=key)
        v = ws.cell(row=row, column=2, value=val)
        bg = LIGHT if shade else WHITE
        for c in (k, v):
            c.fill = fill(bg)
            c.border = border
            c.alignment = left()
        k.font = Font(name="Calibri", bold=True, size=9, color="475569")
        v.font = Font(name="Calibri", size=9, color="0F172A")
        ws.row_dimensions[row].height = 18

    # Executive Summary
    sec(ws1, 4, "EXECUTIVE SUMMARY", NAVY)
    ws1.merge_cells("A5:B5")
    ws1["A5"] = executive_summary
    ws1["A5"].font = Font(name="Calibri", size=9, color="1E293B")
    ws1["A5"].fill = fill(LIGHT)
    ws1["A5"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws1["A5"].border = border
    ws1.row_dimensions[5].height = 52

    ws1.append([])

    # Procurement Details
    sec(ws1, 7, "PROCUREMENT DETAILS", BLUE)
    rows = [
        ("Purchase Order ID", purchase_order_id),
        ("Supplier",          supplier),
        ("Estimated Cost",    _fmt_cost(estimated_cost)),
        ("Approval Status",   approval_status),
        ("Execution Status",  execution_status),
    ]
    for i, (k, v) in enumerate(rows):
        kv(ws1, 8 + i, k, v, shade=(i % 2 == 0))

    ws1.append([])

    # Status Overview
    sec(ws1, 14, "STATUS OVERVIEW", TEAL)
    status_rows = [
        ("Approval",  approval_status,  GREEN if approval_status == "APPROVED" else ORANGE),
        ("Execution", execution_status, GREEN if execution_status == "SUCCESS" else RED),
        ("Dashboard", "UPDATED",        BLUE),
    ]
    for i, (label, val, color) in enumerate(status_rows):
        r = 15 + i
        k = ws1.cell(row=r, column=1, value=label)
        v = ws1.cell(row=r, column=2, value=val)
        k.font = Font(name="Calibri", bold=True, size=9, color="475569")
        k.fill = fill(LIGHT if i % 2 == 0 else WHITE)
        k.border = border
        k.alignment = left()
        v.font = Font(name="Calibri", bold=True, size=9, color=color)
        v.fill = fill(LIGHT if i % 2 == 0 else WHITE)
        v.border = border
        v.alignment = left()
        ws1.row_dimensions[r].height = 18

    # ── Sheet 2: Incident Timeline ───────────────────────────────────────────
    ws2 = wb.create_sheet("Incident Timeline")
    ws2.sheet_view.showGridLines = False
    for col, width in zip("ABCD", [8, 36, 16, 26]):
        ws2.column_dimensions[col].width = width

    ws2.merge_cells("A1:D1")
    ws2["A1"] = "Incident Timeline"
    ws2["A1"].font = Font(name="Calibri", bold=True, size=13, color=WHITE)
    ws2["A1"].fill = fill(PURPLE)
    ws2["A1"].alignment = center()
    ws2.row_dimensions[1].height = 28

    headers = ["Step", "Event", "Status", "Timestamp"]
    for col, h in enumerate(headers, 1):
        c = ws2.cell(row=2, column=col, value=h)
        c.font = hdr_font(WHITE, 9)
        c.fill = fill("4C1D95")
        c.alignment = center()
        c.border = border
    ws2.row_dimensions[2].height = 20

    for i, row in enumerate(timeline):
        r = 3 + i
        shade = LIGHT if i % 2 == 0 else WHITE
        s = row["status"]
        status_color = GREEN if s == "COMPLETED" else (ORANGE if s == "APPROVED" else RED)
        vals = [row["step"], row["event"], row["status"], row["timestamp"]]
        for col, val in enumerate(vals, 1):
            c = ws2.cell(row=r, column=col, value=val)
            c.fill = fill(shade)
            c.border = border
            c.alignment = center() if col != 2 else left()
            c.font = Font(name="Calibri", size=9,
                          color=status_color if col == 3 else "0F172A",
                          bold=(col == 3))
        ws2.row_dimensions[r].height = 18

    # ── Sheet 3: Business Impact ─────────────────────────────────────────────
    ws3 = wb.create_sheet("Business Impact")
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 16
    ws3.column_dimensions["C"].width = 44

    ws3.merge_cells("A1:C1")
    ws3["A1"] = "Business Impact Summary"
    ws3["A1"].font = Font(name="Calibri", bold=True, size=13, color=WHITE)
    ws3["A1"].fill = fill("0F766E")
    ws3["A1"].alignment = center()
    ws3.row_dimensions[1].height = 28

    for col, h in enumerate(["Impact Area", "Value", "Detail"], 1):
        c = ws3.cell(row=2, column=col, value=h)
        c.font = hdr_font(WHITE, 9)
        c.fill = fill("134E4A")
        c.alignment = center()
        c.border = border
    ws3.row_dimensions[2].height = 20

    impact_data = [
        ("Production Impact",  "Moderate",  "2 SKUs halted for 4 hrs — resumed after vendor switch"),
        ("Revenue Impact",     "$284,000",  "Potential revenue loss averted by emergency procurement"),
        ("Customer Impact",    "Minimal",   "3 delayed orders — customers notified proactively"),
        ("Financial Impact",   "$62,000",   "Cloud overspend contained. Recovery plan in progress."),
    ]
    for i, (area, val, detail) in enumerate(impact_data):
        r = 3 + i
        shade = LIGHT if i % 2 == 0 else WHITE
        for col, v in enumerate([area, val, detail], 1):
            c = ws3.cell(row=r, column=col, value=v)
            c.fill = fill(shade)
            c.border = border
            c.font = Font(name="Calibri", size=9, color="0F172A", bold=(col == 1))
            c.alignment = left()
        ws3.row_dimensions[r].height = 18

    filename = f"executive_report_{incident_id}_{uuid.uuid4().hex[:6]}.xlsx"
    filepath = _REPORTS_DIR / filename
    wb.save(str(filepath))
    return {"excel_file": filename, "excel_path": str(filepath), "status": "GENERATED"}


def prepare_dashboard_summary(incident_id: str, purchase_order_id: str, supplier: str,
                                estimated_cost: float, execution_status: str,
                                approval_status: str) -> dict:
    return {
        "incident_id": incident_id,
        "purchase_order_id": purchase_order_id,
        "supplier": supplier,
        "estimated_cost": estimated_cost,
        "execution_status": execution_status,
        "approval_status": approval_status,
        "dashboard_status": "UPDATED",
        "updated_at": _ts(),
    }


def prepare_slack_notification(incident_id: str, purchase_order_id: str, supplier: str,
                                 estimated_cost: float, execution_status: str) -> dict:
    icon = "✅" if execution_status == "SUCCESS" else "❌"
    return {
        "channel": "#procurement-alerts",
        "text": f"{icon} *Procurement Complete* — Incident `{incident_id}`",
        "attachments": [{
            "color": "#22c55e" if execution_status == "SUCCESS" else "#ef4444",
            "fields": [
                {"title": "Purchase Order", "value": purchase_order_id, "short": True},
                {"title": "Supplier",       "value": supplier,          "short": True},
                {"title": "Estimated Cost", "value": _fmt_cost(estimated_cost), "short": True},
                {"title": "Status",         "value": execution_status,  "short": True},
            ],
            "footer": "Enterprise AI Hub",
            "ts": _ts(),
        }],
    }


def prepare_whatsapp_notification(incident_id: str, purchase_order_id: str, supplier: str,
                                    estimated_cost: float, execution_status: str) -> dict:
    icon = "✅" if execution_status == "SUCCESS" else "❌"
    return {
        "to": "+1-ENTERPRISE-OPS",
        "type": "text",
        "text": {
            "body": (
                f"{icon} *Enterprise AI Hub Alert*\n"
                f"Incident: {incident_id}\n"
                f"PO: {purchase_order_id}\n"
                f"Supplier: {supplier}\n"
                f"Cost: {_fmt_cost(estimated_cost)}\n"
                f"Status: {execution_status}"
            )
        },
    }


def archive_report(incident_id: str, report_file: str, executive_summary: str,
                    execution_status: str) -> dict:
    entry = {
        "archive_id": f"ARC-{uuid.uuid4().hex[:8].upper()}",
        "incident_id": incident_id,
        "report_file": report_file,
        "executive_summary": executive_summary[:200],
        "execution_status": execution_status,
        "archived_at": _ts(),
    }
    _report_archive.append(entry)
    return entry


def get_report_archive() -> list:
    return _report_archive
